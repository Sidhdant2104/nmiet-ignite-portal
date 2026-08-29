"""Authenticated internal administration endpoints.

Provision the first organizer with ADMIN_BOOTSTRAP_EMAIL and a bcrypt
ADMIN_BOOTSTRAP_PASSWORD_HASH. Plain-text passwords are never accepted from
environment configuration or written to the database.
"""
from datetime import datetime, timedelta, timezone
from typing import Literal, Optional
from time import perf_counter
import asyncio
import os
import re

import bcrypt
import jwt
from bson import ObjectId
from bson.errors import InvalidId
from fastapi import APIRouter, BackgroundTasks, Cookie, Depends, Header, HTTPException, Request, Response
from fastapi.responses import StreamingResponse
from io import BytesIO, StringIO
import csv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from pydantic import BaseModel, EmailStr, Field

from app.config import ADMIN_BOOTSTRAP_EMAIL, ADMIN_BOOTSTRAP_PASSWORD_HASH, ADMIN_JWT_SECRET, PORTAL_URL
from copy import deepcopy
from app.mongodb import admin_users_collection, announcement_collection, audit_collection, registration_collection, settings_collection
from app.routes.ppt import STATUS as PPT_STATUSES, log_email
from app.services.storage import create_signed_download, create_signed_preview, download_file as storage_download
from zipfile import ZIP_DEFLATED, ZipFile

router = APIRouter(prefix="/admin", tags=["Administration"])
Role = Literal["super_admin", "faculty", "student_spoc", "student_coordinator"]
COOKIE_NAME = "nmiet_admin_session"

ROLE_PERMISSIONS = {
    "super_admin": {"manage_registrations", "delete_registrations", "manage_announcements", "send_email", "manage_users", "export", "view_dashboard", "view_activity", "view_ppt", "review_ppt", "download_ppt", "manage_evaluation"},
    "faculty": {"manage_registrations", "export", "view_dashboard", "view_ppt", "review_ppt", "download_ppt", "manage_evaluation"},
    "student_spoc": {"manage_registrations", "view_dashboard", "export", "view_ppt", "manage_evaluation"},
    "student_coordinator": {"view_dashboard", "view_ppt"},
}

class LoginPayload(BaseModel):
    email: EmailStr
    password: str = Field(min_length=8, max_length=256)

class RegistrationPatch(BaseModel):
    team: Optional[dict] = None
    leader: Optional[dict] = None
    members: Optional[list[dict]] = None
    mentor: Optional[dict] = None
    status: Optional[Literal["Registered", "PPT Submitted", "Under Review", "Shortlisted", "Rejected", "Qualified"]] = None
    remarks: Optional[str] = Field(default=None, max_length=1000)

class AnnouncementPayload(BaseModel):
    title: str = Field(min_length=3, max_length=140)
    body: str = Field(min_length=1, max_length=5000)
    tag: str = Field(default="Update", min_length=1, max_length=48)
    is_pinned: bool = False
    is_published: bool = True
    scheduled_for: Optional[datetime] = None
    expires_at: Optional[datetime] = None

class AdminUserPayload(BaseModel):
    name: str = Field(min_length=2, max_length=80)
    email: EmailStr
    role: Role
    password: Optional[str] = Field(default=None, min_length=8, max_length=256)
    is_active: bool = True

class BulkRegistrationPayload(BaseModel):
    ids: list[str] = Field(min_length=1, max_length=500)
    action: Literal["delete", "restore", "status"]
    status: Optional[Literal["Registered", "PPT Submitted", "Under Review", "Shortlisted", "Rejected", "Qualified"]] = None

class InvitePayload(BaseModel):
    name: str = Field(min_length=2, max_length=80)
    email: EmailStr
    role: Role

class PptReviewPayload(BaseModel):
    status: Literal["PPT Submitted", "Under Review", "Revision Requested", "Approved", "Rejected", "Qualified"]
    reviewer_remarks: str = Field(default="", max_length=3000)
    internal_notes: str = Field(default="", max_length=3000)

class RegistrationControlPayload(BaseModel):
    is_open: bool

def _secret() -> str:
    if not ADMIN_JWT_SECRET or len(ADMIN_JWT_SECRET) < 32:
        raise HTTPException(status_code=503, detail="Admin authentication is not configured.")
    return ADMIN_JWT_SECRET

async def _bootstrap_user():
    if not ADMIN_BOOTSTRAP_EMAIL or not ADMIN_BOOTSTRAP_PASSWORD_HASH:
        return None
    email = ADMIN_BOOTSTRAP_EMAIL.lower()
    user = await admin_users_collection.find_one({"email": email})
    if not user:
        await admin_users_collection.insert_one({"email": email, "name": "Super Admin", "role": "super_admin", "password_hash": ADMIN_BOOTSTRAP_PASSWORD_HASH, "created_at": datetime.now(timezone.utc)})
        user = await admin_users_collection.find_one({"email": email})
    return user

async def current_admin(
    authorization: Optional[str] = Header(default=None),
    session: Optional[str] = Cookie(default=None, alias=COOKIE_NAME),
):
    started_at = perf_counter()
    # Prefer the explicit Bearer token used by the frontend. The cookie fallback
    # keeps existing same-site/direct-download admin links working during rollout.
    if authorization is not None:
        scheme, _, token = authorization.partition(" ")
        if scheme.lower() != "bearer" or not token:
            raise HTTPException(status_code=401, detail="Authentication required.")
    else:
        token = session
    if not token:
        raise HTTPException(status_code=401, detail="Authentication required.")
    try:
        payload = jwt.decode(token, _secret(), algorithms=["HS256"])
        user_id = ObjectId(payload["sub"])
    except (jwt.PyJWTError, InvalidId, KeyError, TypeError, ValueError):
        raise HTTPException(status_code=401, detail="Session is invalid or expired.")
    print(f"ADMIN AUTH TOKEN VERIFY: {perf_counter() - started_at:.3f}s")
    user = await admin_users_collection.find_one({"_id": user_id, "is_active": {"$ne": False}})
    print(f"ADMIN AUTH MONGO LOOKUP: {perf_counter() - started_at:.3f}s total")
    if not user:
        raise HTTPException(status_code=401, detail="Session is invalid.")
    user["_id"] = str(user["_id"])
    return user

def require(permission: str):
    async def guard(user=Depends(current_admin)):
        if permission not in ROLE_PERMISSIONS.get(user["role"], set()):
            raise HTTPException(status_code=403, detail="You do not have permission for this action.")
        return user
    return guard

async def csrf_guard(request: Request):
    """Require browser mutations to originate from the approved portal origins."""
    origin = request.headers.get("origin")
    allowed = {
    "http://localhost:5173",
    "http://127.0.0.1:5173",
    "http://localhost:8080",
    "http://127.0.0.1:8080",
    "http://localhost:8000",
    "http://127.0.0.1:8000",
    "https://nmietsihportal.vercel.app",
}
    if origin and origin not in allowed:
        raise HTTPException(status_code=403, detail="Cross-site request blocked.")

async def audit(user: dict, action: str, registration_id: Optional[str] = None, detail: Optional[str] = None):
    await audit_collection.insert_one({"admin_id": user["_id"], "admin_name": user.get("name", user["email"]), "action": action, "registration_id": registration_id, "detail": detail, "timestamp": datetime.now(timezone.utc)})

def ppt_scope(user: dict):
    if user["role"] == "super_admin": return {"isDeleted": {"$ne": True}}
    # Faculty and coordinators are restricted to explicitly assigned teams.
    if user["role"] == "faculty": return {"isDeleted": {"$ne": True}, "mentor.email": user["email"]}
    return {"isDeleted": {"$ne": True}, "$or": [{"coordinator_email": user["email"]}, {"team.coordinator_email": user["email"]}]}

def serialize(item: dict):
    safe = deepcopy(item); safe["_id"] = str(safe["_id"])
    for upload in safe.get("ppt", {}).get("history", []):
        upload.pop("storage_key", None)
        upload.get("storage", {}).pop("original_key", None)
        upload.get("storage", {}).pop("preview_key", None)
    current = safe.get("ppt", {}).get("current")
    if current:
        current.pop("storage_key", None)
        current.get("storage", {}).pop("original_key", None)
        current.get("storage", {}).pop("preview_key", None)
    return safe

def ppt_original_key(upload: dict) -> Optional[str]:
    """Read both the legacy and current private-storage document shapes."""
    return upload.get("storage_key") or upload.get("storage", {}).get("original_key")

def ppt_preview_key(upload: dict) -> Optional[str]:
    return upload.get("preview_key") or upload.get("storage", {}).get("preview_key") or (
        ppt_original_key(upload) if upload.get("format") == "pdf" or str(upload.get("original_filename") or upload.get("file_name") or "").lower().endswith(".pdf") else None
    )

def upload_filename(upload: dict) -> str:
    return upload.get("original_filename") or upload.get("file_name") or "presentation"

def zip_name(value: str, fallback: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._ -]+", "_", value or "").strip(" .")
    return cleaned or fallback

async def stream_ppt_zip(items: list[dict], filename: str, user: dict, root: Optional[str] = None):
    """Build an admin-only archive from private storage without leaking object keys."""
    output = BytesIO()
    try:
        with ZipFile(output, "w", ZIP_DEFLATED) as archive:
            for item in items:
                current = item.get("ppt", {}).get("current") or {}
                key = ppt_original_key(current)
                if not key:
                    continue
                try:
                    content = storage_download(key)
                except RuntimeError:
                    # One unavailable object should not prevent recovery of every other submission.
                    continue
                team = zip_name(item.get("team", {}).get("teamName", ""), "Untitled Team")
                theme = zip_name(item.get("team", {}).get("theme", ""), "Unassigned")
                name = zip_name(upload_filename(current), "presentation")
                prefix = f"{root}/" if root else ""
                archive.writestr(f"{prefix}{theme}/{team}/{name}", content)
    except RuntimeError as error:
        raise HTTPException(502, "Unable to retrieve presentation. Please try again.") from error
    output.seek(0)
    await audit(user, "Downloaded PPT ZIP", detail=filename)
    return StreamingResponse(output, media_type="application/zip", headers={"Content-Disposition": f'attachment; filename="{filename}"'})

def registration_query(search: Optional[str] = None, status: Optional[str] = None, theme: Optional[str] = None, category: Optional[str] = None, date_from: Optional[datetime] = None, date_to: Optional[datetime] = None, include_deleted: bool = False):
    query = {} if include_deleted else {"isDeleted": {"$ne": True}}
    if status:
        query["status"] = status
    if theme:
        query["team.theme"] = theme
    if category:
        query["team.category"] = category
    if date_from or date_to:
        query["created_at"] = {**({"$gte": date_from} if date_from else {}), **({"$lte": date_to} if date_to else {})}
    if search:
        query["$or"] = [
            {"team.teamName": {"$regex": search, "$options": "i"}},
            {"leader.name": {"$regex": search, "$options": "i"}},
            {"leader.email": {"$regex": search, "$options": "i"}},
            {"team.psId": {"$regex": search, "$options": "i"}},
            {"registration_id": {"$regex": search, "$options": "i"}},
            {"leader.college": {"$regex": search, "$options": "i"}},
            {"team.college": {"$regex": search, "$options": "i"}},
        ]
    return query

def safe_user(user: dict):
    return {"_id": str(user["_id"]), "name": user.get("name"), "email": user["email"], "role": user["role"], "is_active": user.get("is_active", True), "created_at": user.get("created_at")}

@router.get("/users")
async def users(search: Optional[str] = None, user=Depends(require("manage_users"))):
    query = {"is_deleted": {"$ne": True}}
    if search: query["$or"] = [{"name": {"$regex": search, "$options": "i"}}, {"email": {"$regex": search, "$options": "i"}}, {"role": {"$regex": search, "$options": "i"}}]
    result=[]
    async for item in admin_users_collection.find(query).sort("created_at", -1): result.append(safe_user(item))
    return {"data": result}

@router.post("/users", dependencies=[Depends(csrf_guard)])
async def create_user(payload: AdminUserPayload, user=Depends(require("manage_users"))):
    if await admin_users_collection.find_one({"email": payload.email.lower(), "is_deleted": {"$ne": True}}): raise HTTPException(409, "An admin with this email already exists.")
    if not payload.password: raise HTTPException(422, "Use the invitation workflow to set the password.")
    item=payload.model_dump(); item["email"]=item["email"].lower(); item["password_hash"]=bcrypt.hashpw(item.pop("password").encode(),bcrypt.gensalt()).decode(); item["created_at"]=datetime.now(timezone.utc)
    result=await admin_users_collection.insert_one(item); await audit(user,"Created Admin",detail=item["email"]); return {"id":str(result.inserted_id)}

@router.patch("/users/{user_id}", dependencies=[Depends(csrf_guard)])
async def update_user(user_id: str, payload: AdminUserPayload, user=Depends(require("manage_users"))):
    data=payload.model_dump(exclude={"password"});
    if payload.password: data["password_hash"]=bcrypt.hashpw(payload.password.encode(),bcrypt.gensalt()).decode()
    result=await admin_users_collection.update_one({"_id":ObjectId(user_id),"is_deleted":{"$ne":True}},{"$set":data})
    if not result.matched_count: raise HTTPException(404,"Admin not found.")
    await audit(user,"Changed role" if "role" in data else "Updated Admin",detail=payload.email); return {"success":True}

@router.delete("/users/{user_id}", dependencies=[Depends(csrf_guard)])
async def delete_user(user_id: str, user=Depends(require("manage_users"))):
    if user_id==user["_id"]: raise HTTPException(400,"You cannot delete your own account.")
    result=await admin_users_collection.update_one({"_id":ObjectId(user_id)},{"$set":{"is_deleted":True,"is_active":False,"deleted_at":datetime.now(timezone.utc)}})
    if not result.matched_count: raise HTTPException(404,"Admin not found.")
    await audit(user,"Deleted Admin",detail=user_id); return {"success":True}

@router.post("/users/invite", dependencies=[Depends(csrf_guard)])
async def invite_user(payload: InvitePayload, request: Request, user=Depends(require("manage_users"))):
    if await admin_users_collection.find_one({"email": payload.email.lower(), "is_deleted": {"$ne": True}}): raise HTTPException(409,"An admin with this email already exists.")
    token=jwt.encode({"email":payload.email.lower(),"name":payload.name,"role":payload.role,"purpose":"admin_invite","exp":datetime.now(timezone.utc)+timedelta(hours=24)},_secret(),algorithm="HS256")
    await admin_users_collection.insert_one({"name":payload.name,"email":payload.email.lower(),"role":payload.role,"is_active":False,"invite_token":token,"invite_expires_at":datetime.now(timezone.utc)+timedelta(hours=24),"created_at":datetime.now(timezone.utc)})
    await audit(user,"Invited Admin",detail=payload.email)
    return {"invite_url":f"{request.headers.get('origin','')}/admin/accept-invite?token={token}","expires_in_hours":24}

@router.post("/registrations/bulk", dependencies=[Depends(csrf_guard)])
async def bulk_registrations(payload: BulkRegistrationPayload, user=Depends(require("manage_registrations"))):
    ids=[ObjectId(value) for value in payload.ids]
    if payload.action=="delete":
        if "delete_registrations" not in ROLE_PERMISSIONS.get(user["role"],set()): raise HTTPException(403,"You do not have permission to delete registrations.")
        await registration_collection.update_many({"_id":{"$in":ids}},{"$set":{"isDeleted":True,"deleted_at":datetime.now(timezone.utc),"deleted_by":user["_id"]}})
    elif payload.action=="restore": await registration_collection.update_many({"_id":{"$in":ids}},{"$set":{"isDeleted":False}})
    elif payload.action=="status":
        if not payload.status: raise HTTPException(422,"Status is required.")
        await registration_collection.update_many({"_id":{"$in":ids}},{"$set":{"status":payload.status,"updated_at":datetime.now(timezone.utc)}})
    await audit(user,f"Bulk {payload.action.title()} Registrations",detail=f"{len(ids)} registrations")
    return {"success":True,"count":len(ids)}

@router.get("/registrations/export")
async def export_registrations(format: Literal["csv","xlsx"]="csv", search: Optional[str] = None, status: Optional[str] = None, theme: Optional[str] = None, category: Optional[str] = None, date_from: Optional[datetime] = None, date_to: Optional[datetime] = None, include_deleted: bool = False, sort_by: str = "created_at", sort_order: int = -1, user=Depends(require("export"))):
    headers = ["Registration ID", "Registration Date", "Status", "Team Name", "PS ID", "Theme", "Category", "Leader Name", "Leader Email", "Leader Phone", "Leader Department", "Leader Year", "Faculty Name", "Faculty Email", "Faculty Phone"]
    for index in range(1, 7):
        headers.extend([f"Member {index} Name", f"Member {index} Email", f"Member {index} Phone", f"Member {index} Department", f"Member {index} Year"])
    headers.append("Remarks")
    safe_sort = sort_by if sort_by in {"created_at", "status", "team.teamName", "team.theme"} else "created_at"
    registrations = []
    async for registration in registration_collection.find(registration_query(search, status, theme, category, date_from, date_to, include_deleted)).sort(safe_sort, 1 if sort_order == 1 else -1):
        registrations.append(registration)

    def created_at(value, for_excel=False):
        if not isinstance(value, datetime):
            return value or ""
        normalized = value.astimezone(timezone.utc).replace(tzinfo=None) if value.tzinfo else value
        return normalized if for_excel else normalized.strftime("%Y-%m-%d %H:%M UTC")

    def registration_row(registration, for_excel=False):
        team, leader, mentor = registration.get("team", {}), registration.get("leader", {}), registration.get("mentor") or {}
        row = [registration.get("registration_id") or str(registration.get("_id", "")), created_at(registration.get("created_at"), for_excel), registration.get("status", "Registered"), team.get("teamName", ""), team.get("psId", ""), team.get("theme", ""), team.get("category", ""), leader.get("name", ""), leader.get("email", ""), leader.get("mobile", ""), leader.get("department", ""), leader.get("year", ""), mentor.get("name", ""), mentor.get("email", ""), mentor.get("mobile", "")]
        for member in registration.get("members", [])[:6]:
            row.extend([member.get("name", ""), member.get("email", ""), member.get("mobile", ""), member.get("department", ""), member.get("year", "")])
        row.extend([""] * (30 - len(row) + 15))
        row.append(registration.get("remarks", ""))
        return row

    def apply_sheet_style(sheet, status_column=None, date_column=None):
        thin = Side(style="thin", color="D1D5DB")
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="F97316")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
        for row_index, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            fill = PatternFill("solid", fgColor="FFF7ED") if row_index % 2 == 0 else None
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
                if fill:
                    cell.fill = fill
        if status_column:
            for cell in sheet[status_column][1:]:
                cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
        if date_column:
            for cell in sheet[date_column][1:]:
                cell.number_format = "yyyy-mm-dd hh:mm"
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column in sheet.columns:
            sheet.column_dimensions[column[0].column_letter].width = min(max(len(str(cell.value or "")) for cell in column) + 2, 36)

    filename=f"registrations-{datetime.now().date()}.{format}"
    await audit(user,f"Exported {format.upper()}",detail=filename)
    if format=="csv":
        stream=StringIO(); writer=csv.writer(stream); writer.writerow(headers); writer.writerows(registration_row(registration) for registration in registrations); return Response("\ufeff" + stream.getvalue(),media_type="text/csv; charset=utf-8",headers={"Content-Disposition":f'attachment; filename="{filename}"'})

    workbook = Workbook()
    registrations_sheet = workbook.active
    registrations_sheet.title = "Registrations"
    registrations_sheet.append(headers)
    for registration in registrations:
        registrations_sheet.append(registration_row(registration, True))
    apply_sheet_style(registrations_sheet, status_column="C", date_column="B")

    members_sheet = workbook.create_sheet("Team Members")
    members_sheet.append(["Registration ID", "Team Name", "PS ID", "Role", "Name", "Email", "Phone", "Department", "Year"])
    for registration in registrations:
        team, leader = registration.get("team", {}), registration.get("leader", {})
        base = [registration.get("registration_id") or str(registration.get("_id", "")), team.get("teamName", ""), team.get("psId", "")]
        members_sheet.append(base + ["Leader", leader.get("name", ""), leader.get("email", ""), leader.get("mobile", ""), leader.get("department", ""), leader.get("year", "")])
        for index, member in enumerate(registration.get("members", []), start=1):
            members_sheet.append(base + [f"Member {index}", member.get("name", ""), member.get("email", ""), member.get("mobile", ""), member.get("department", ""), member.get("year", "")])
    apply_sheet_style(members_sheet)

    summary_sheet = workbook.create_sheet("Summary")
    active = [registration for registration in registrations if not registration.get("isDeleted", False)]
    summary_sheet.append(["Metric", "Value"])
    summary_values = [("Total Registrations", len(active)), ("Software Teams", sum(item.get("team", {}).get("category") == "Software" for item in active)), ("Hardware Teams", sum(item.get("team", {}).get("category") == "Hardware" for item in active))]
    for status_name in ["Registered", "PPT Submitted", "Under Review", "Shortlisted", "Qualified", "Rejected"]:
        summary_values.append((status_name, sum(item.get("status", "Registered") == status_name for item in active)))
    summary_values.append(("Deleted", sum(item.get("isDeleted", False) for item in registrations)))
    for row in summary_values:
        summary_sheet.append(row)
    apply_sheet_style(summary_sheet)

    out=BytesIO(); workbook.save(out); out.seek(0); return StreamingResponse(out,media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",headers={"Content-Disposition":f'attachment; filename="{filename}"'})

@router.post("/auth/login", dependencies=[Depends(csrf_guard)])
async def login(data: LoginPayload, response: Response, background_tasks: BackgroundTasks):
    started_at = perf_counter()
    print("ADMIN LOGIN START")
    step_started_at = perf_counter()
    await _bootstrap_user()
    print(f"ADMIN LOGIN BOOTSTRAP: {perf_counter() - step_started_at:.3f}s")
    step_started_at = perf_counter()
    user = await admin_users_collection.find_one({"email": data.email.lower(), "is_active": {"$ne": False}})
    print(f"ADMIN LOGIN MONGO LOOKUP: {perf_counter() - step_started_at:.3f}s")
    step_started_at = perf_counter()
    if not user or not bcrypt.checkpw(data.password.encode(), user["password_hash"].encode()):
        raise HTTPException(status_code=401, detail="Invalid email or password.")
    print(f"ADMIN LOGIN PASSWORD VERIFY: {perf_counter() - step_started_at:.3f}s")
    step_started_at = perf_counter()
    token = jwt.encode({"sub": str(user["_id"]), "role": user["role"], "exp": datetime.now(timezone.utc) + timedelta(hours=8)}, _secret(), algorithm="HS256")
    production = os.getenv("ENVIRONMENT") == "production"
    response.set_cookie(COOKIE_NAME, token, httponly=True, secure=production, samesite="none" if production else "lax", max_age=28800, path="/")
    background_tasks.add_task(audit, {**user, "_id": str(user["_id"])}, "Signed in")
    print(f"ADMIN LOGIN TOKEN AND AUDIT QUEUE: {perf_counter() - step_started_at:.3f}s")
    print(f"ADMIN LOGIN RESPONSE: {perf_counter() - started_at:.3f}s total")
    return {"token": token, "user": {"name": user.get("name", "Administrator"), "email": user["email"], "role": user["role"]}}

@router.post("/auth/logout", status_code=204, dependencies=[Depends(csrf_guard)])
async def logout(response: Response):
    response.delete_cookie(COOKIE_NAME, path="/")

@router.get("/auth/me")
async def me(user=Depends(current_admin)):
    started_at = perf_counter()
    response = {"name": user.get("name", "Administrator"), "email": user["email"], "role": user["role"]}
    print(f"ADMIN AUTH ME RESPONSE: {perf_counter() - started_at:.3f}s (authentication timed by dependency)")
    return response

@router.get("/dashboard")
async def dashboard(user=Depends(current_admin)):
    started_at = perf_counter()
    print("ADMIN DASHBOARD START")
    base = {"isDeleted": {"$ne": True}}
    dashboard_data, activity, control = await asyncio.gather(
        registration_collection.aggregate([
        {"$match": base},
        {"$facet": {
            "metrics": [{"$group": {"_id": None, "total": {"$sum": 1}, "software": {"$sum": {"$cond": [{"$eq": ["$team.category", "Software"]}, 1, 0]}}, "hardware": {"$sum": {"$cond": [{"$eq": ["$team.category", "Hardware"]}, 1, 0]}}, "ppt_submitted": {"$sum": {"$cond": [{"$ne": [{"$type": "$ppt.current"}, "missing"]}, 1, 0]}}, "awaiting_ppt": {"$sum": {"$cond": [{"$eq": [{"$type": "$ppt.current"}, "missing"]}, 1, 0]}}, "under_review": {"$sum": {"$cond": [{"$eq": ["$ppt.current.status", "Under Review"]}, 1, 0]}}, "approved": {"$sum": {"$cond": [{"$eq": ["$ppt.current.status", "Approved"]}, 1, 0]}}, "revision_requested": {"$sum": {"$cond": [{"$eq": ["$ppt.current.status", "Revision Requested"]}, 1, 0]}}, "pending_ppt": {"$sum": {"$cond": [{"$eq": ["$status", "Registered"]}, 1, 0]}}, "shortlisted": {"$sum": {"$cond": [{"$eq": ["$status", "Shortlisted"]}, 1, 0]}}, "rejected": {"$sum": {"$cond": [{"$eq": ["$status", "Rejected"]}, 1, 0]}}, "qualified": {"$sum": {"$cond": [{"$eq": ["$status", "Qualified"]}, 1, 0]}}}}],
            "status_distribution": [{"$group": {"_id": "$status", "count": {"$sum": 1}}}],
            "theme_distribution": [{"$group": {"_id": "$team.theme", "count": {"$sum": 1}}}, {"$sort": {"count": -1}}, {"$limit": 8}],
            "latest": [{"$sort": {"created_at": -1}}, {"$limit": 6}, {"$project": {"registration_id": 1, "team.teamName": 1, "status": 1, "created_at": 1}}],
        }},
        ]).to_list(length=1),
        audit_collection.find({}, {"admin_name": 1, "action": 1, "registration_id": 1, "detail": 1, "timestamp": 1}).sort("timestamp", -1).limit(6).to_list(length=6),
        settings_collection.find_one({"key": "registration_control"}),
    )
    print(f"ADMIN DASHBOARD MONGO AGGREGATION: {perf_counter() - started_at:.3f}s")
    result = dashboard_data[0] if dashboard_data else {}
    metrics = (result.get("metrics") or [{"total": 0}])[0]
    statuses = ["Registered", "PPT Submitted", "Under Review", "Shortlisted", "Rejected", "Qualified"]
    status_distribution = {status: 0 for status in statuses}
    status_distribution.update({item.get("_id") or "Registered": item["count"] for item in result.get("status_distribution", [])})
    latest = result.get("latest", [])
    for item in latest:
        item["_id"] = str(item["_id"])
    theme_distribution = [{"theme": item["_id"] or "Unspecified", "count": item["count"]} for item in result.get("theme_distribution", [])]
    for item in activity:
        item["_id"] = str(item["_id"])
    print(f"ADMIN DASHBOARD RESPONSE: {perf_counter() - started_at:.3f}s total")
    return {"registration_open": True if not control else control.get("is_open", True), "metrics": metrics, "status_distribution": status_distribution, "theme_distribution": theme_distribution, "latest": latest, "activity": activity}

@router.get("/registration-control")
async def get_registration_control(user=Depends(current_admin)):
    control = await settings_collection.find_one({"key": "registration_control"})
    return {"is_open": True if not control else control.get("is_open", True)}

@router.put("/registration-control", dependencies=[Depends(csrf_guard)])
async def set_registration_control(payload: RegistrationControlPayload, user=Depends(require("manage_users"))):
    if user["role"] != "super_admin": raise HTTPException(403, "Only Super Admin can change registration availability.")
    await settings_collection.update_one({"key": "registration_control"}, {"$set": {"is_open": payload.is_open, "updated_at": datetime.now(timezone.utc), "updated_by": user["_id"]}}, upsert=True)
    await audit(user, "Registrations opened" if payload.is_open else "Registrations closed", detail=f"Registrations {'opened' if payload.is_open else 'closed'} by Super Admin")
    return {"is_open": payload.is_open}

@router.get("/ppt")
@router.get("/ppt/submissions")
async def ppt_submissions(user=Depends(require("view_ppt"))):
    results = []
    projection = {
        "registration_id": 1,
        "team.teamName": 1,
        "team.psId": 1,
        "team.theme": 1,
        "team.category": 1,
        "team.college": 1,
        "leader.name": 1,
        "leader.email": 1,
        "leader.college": 1,
        "mentor.name": 1,
        "ppt.current": 1,
    }
    async for item in registration_collection.find(ppt_scope(user), projection).sort("ppt.current.uploaded_at", -1): results.append(serialize(item))
    return {"data": results}

@router.get("/ppt/themes")
async def ppt_themes(user=Depends(require("view_ppt"))):
    """Metadata-only theme summary; files are fetched only on review/download."""
    groups: dict[str, dict] = {}
    async for item in registration_collection.find(ppt_scope(user), {"team.theme": 1, "ppt.current.status": 1}):
        theme = item.get("team", {}).get("theme") or "Unassigned"
        group = groups.setdefault(theme, {"theme": theme, "total_teams": 0, "ppt_submitted": 0, "pending_review": 0, "approved": 0, "revision_requested": 0, "rejected": 0})
        group["total_teams"] += 1
        upload = item.get("ppt", {}).get("current")
        if not upload:
            continue
        group["ppt_submitted"] += 1
        status = upload.get("status", "PPT Submitted")
        if status in {"PPT Submitted", "Under Review"}: group["pending_review"] += 1
        elif status == "Approved": group["approved"] += 1
        elif status == "Revision Requested": group["revision_requested"] += 1
        elif status == "Rejected": group["rejected"] += 1
    return {"data": sorted(groups.values(), key=lambda value: value["theme"].lower())}

@router.get("/ppt/themes/{theme}/download")
async def download_theme_ppts(theme: str, user=Depends(require("download_ppt"))):
    items = []
    theme_query = {"$or": [{"team.theme": {"$exists": False}}, {"team.theme": ""}, {"team.theme": None}]} if theme == "Unassigned" else {"team.theme": theme}
    query = {**ppt_scope(user), **theme_query, "ppt.current": {"$exists": True}}
    async for item in registration_collection.find(query):
        if ppt_original_key(item.get("ppt", {}).get("current") or {}): items.append(item)
    return await stream_ppt_zip(items, f"{zip_name(theme, 'Unassigned')}.zip", user)

@router.get("/ppt/download-all")
async def download_all_ppts(user=Depends(require("download_ppt"))):
    items = []
    query = {**ppt_scope(user), "ppt.current": {"$exists": True}}
    async for item in registration_collection.find(query):
        if ppt_original_key(item.get("ppt", {}).get("current") or {}): items.append(item)
    return await stream_ppt_zip(items, "NMIET_SIH_2026_PPT_SUBMISSIONS.zip", user, "NMIET_SIH_2026_PPT_SUBMISSIONS")

async def _ppt_file(registration_id: str, user: dict):
    try: object_id = ObjectId(registration_id)
    except Exception: raise HTTPException(400, "Invalid registration ID.")
    item = await registration_collection.find_one({"_id": object_id, **ppt_scope(user)})
    current = item.get("ppt", {}).get("current") if item else None
    if not current or not ppt_original_key(current): raise HTTPException(404, "PPT submission not found or not permitted.")
    return item, current

def ppt_version(item: dict, version: Optional[int]) -> dict:
    if version is None:
        return item.get("ppt", {}).get("current") or {}
    for upload in [item.get("ppt", {}).get("current") or {}, *item.get("ppt", {}).get("history", [])]:
        if upload.get("version") == version:
            return upload
    raise HTTPException(404, "PPT version not found.")

@router.get("/ppt/submissions/{registration_id}")
@router.get("/ppt/{registration_id}")
async def ppt_submission(registration_id: str, user=Depends(require("view_ppt"))):
    item, _ = await _ppt_file(registration_id, user)
    return serialize(item)

@router.get("/ppt/{registration_id}/download")
async def ppt_download(registration_id: str, version: Optional[int] = None, user=Depends(require("download_ppt"))):
    item, _ = await _ppt_file(registration_id, user)
    current = ppt_version(item, version)
    try:
        signed_url = create_signed_download(ppt_original_key(current), expires_in=60)
    except RuntimeError as error:
        raise HTTPException(502, "Storage could not generate a download link.") from error
    await audit(user, "Faculty downloaded PPT", registration_id, f"Version {current.get('version')}: {current.get('file_name')}")
    from fastapi.responses import RedirectResponse
    return RedirectResponse(url=signed_url, status_code=302)

@router.get("/ppt/{registration_id}/preview")
async def ppt_preview(registration_id: str, version: Optional[int] = None, user=Depends(require("download_ppt"))):
    item, _ = await _ppt_file(registration_id, user)
    current = ppt_version(item, version)
    key = ppt_preview_key(current)
    if not key: raise HTTPException(422, "Preview unavailable. Download the original presentation instead.")
    try:
        signed_url = create_signed_preview(key, expires_in=60)
    except RuntimeError as error:
        raise HTTPException(502, "Storage could not generate a preview link.") from error
    await audit(user, "Previewed PPT", registration_id, f"Version {current.get('version')}: {current.get('file_name')}")
    from fastapi.responses import RedirectResponse
    return RedirectResponse(url=signed_url, status_code=302)

@router.patch("/ppt/submissions/{registration_id}", dependencies=[Depends(csrf_guard)])
async def review_ppt(registration_id: str, payload: PptReviewPayload, user=Depends(require("review_ppt"))):
    try: object_id = ObjectId(registration_id)
    except Exception: raise HTTPException(400, "Invalid registration ID.")
    item = await registration_collection.find_one({"_id": object_id, **ppt_scope(user)})
    if not item or not item.get("ppt", {}).get("current"): raise HTTPException(404, "PPT submission not found.")
    now = datetime.now(timezone.utc); current = item["ppt"]["current"]
    current.update({"status": payload.status, "reviewer_remarks": payload.reviewer_remarks, "internal_notes": payload.internal_notes, "last_modified": now, "reviewed_by": user["email"]})
    history = item["ppt"].get("history", []);
    if history: history[-1].update({"status": payload.status, "reviewer_remarks": payload.reviewer_remarks})
    await registration_collection.update_one({"_id": object_id}, {"$set": {"ppt.current": current, "ppt.history": history, "status": payload.status, "updated_at": now}})
    await audit(user, "Reviewed PPT", registration_id, f"{payload.status}: {payload.reviewer_remarks[:180]}")
    leader = item.get("leader", {}); team = item.get("team", {})
    await log_email(leader.get("email", ""), f"PPT review update: {payload.status}", f"Hello {leader.get('name', 'Team Leader')},\n\nYour PPT submission for {team.get('teamName', 'your team')} is now: {payload.status}.\nReference ID: {item.get('registration_id')}\n\nReviewer remarks:\n{payload.reviewer_remarks or 'No remarks provided.'}\n\n{PORTAL_URL}/ppt-submission")
    return {"success": True}

@router.get("/registrations")
async def registrations(search: Optional[str] = None, status: Optional[str] = None, theme: Optional[str] = None, category: Optional[str] = None, date_from: Optional[datetime] = None, date_to: Optional[datetime] = None, include_deleted: bool = False, sort_by: str = "created_at", sort_order: int = -1, user=Depends(current_admin)):
    query = registration_query(search, status, theme, category, date_from, date_to, include_deleted)
    safe_sort = sort_by if sort_by in {"created_at", "status", "team.teamName", "team.theme"} else "created_at"
    result = []
    async for item in registration_collection.find(query).sort(safe_sort, 1 if sort_order == 1 else -1): result.append(serialize(item))
    return {"data": result}

@router.get("/registrations/{registration_id}")
async def registration(registration_id: str, include_deleted: bool = False, user=Depends(current_admin)):
    try:
        object_id = ObjectId(registration_id)
    except Exception:
        raise HTTPException(400, "Invalid registration ID.")
    item = await registration_collection.find_one({"_id": object_id, **({} if include_deleted else {"isDeleted": {"$ne": True}})})
    if not item: raise HTTPException(404, "Registration not found.")
    return serialize(item)

@router.patch("/registrations/{registration_id}", dependencies=[Depends(csrf_guard)])
async def update_registration(registration_id: str, payload: RegistrationPatch, user=Depends(require("manage_registrations"))):
    changes = payload.model_dump(exclude_none=True); changes["updated_at"] = datetime.now(timezone.utc)
    old = await registration_collection.find_one({"_id": ObjectId(registration_id), "isDeleted": {"$ne": True}})
    if not old: raise HTTPException(404, "Registration not found.")
    await registration_collection.update_one({"_id": old["_id"]}, {"$set": changes})
    if "status" in changes and changes["status"] != old.get("status"):
        await audit(user, "Changed Status", registration_id, f"{old.get('status', 'Registered')} → {changes['status']}")
    else:
        await audit(user, "Updated Registration", registration_id)
    return {"success": True}

@router.delete("/registrations/{registration_id}", dependencies=[Depends(csrf_guard)])
async def soft_delete(registration_id: str, user=Depends(require("manage_registrations"))):
    result = await registration_collection.update_one({"_id": ObjectId(registration_id), "isDeleted": {"$ne": True}}, {"$set": {"isDeleted": True, "deleted_at": datetime.now(timezone.utc), "deleted_by": user["_id"]}})
    if not result.matched_count: raise HTTPException(404, "Registration not found.")
    await audit(user, "Deleted Registration", registration_id); return {"success": True}

@router.post("/registrations/{registration_id}/restore", dependencies=[Depends(csrf_guard)])
async def restore(registration_id: str, user=Depends(require("manage_registrations"))):
    await registration_collection.update_one({"_id": ObjectId(registration_id), "isDeleted": True}, {"$set": {"isDeleted": False}, "$unset": {"deleted_at": "", "deleted_by": ""}})
    await audit(user, "Restored Registration", registration_id); return {"success": True}

@router.get("/activity")
async def activity(user=Depends(current_admin)):
    items=[]
    async for item in audit_collection.find().sort("timestamp", -1).limit(100): item["_id"] = str(item["_id"]); items.append(item)
    return {"data": items}

@router.get("/announcements")
async def get_announcements(user=Depends(current_admin)):
    items=[]
    async for item in announcement_collection.find().sort("created_at", -1): item["_id"] = str(item["_id"]); items.append(item)
    return {"data": items}

@router.post("/announcements", dependencies=[Depends(csrf_guard)])
async def create_announcement(payload: AnnouncementPayload, user=Depends(require("manage_announcements"))):
    item=payload.model_dump(); item.update({"created_at": datetime.now(timezone.utc), "created_by": user["_id"], "is_archived": False})
    result=await announcement_collection.insert_one(item); await audit(user, "Created announcement", detail=item["title"])
    return {"id": str(result.inserted_id)}

@router.patch("/announcements/{announcement_id}", dependencies=[Depends(csrf_guard)])
async def update_announcement(announcement_id: str, payload: AnnouncementPayload, user=Depends(require("manage_announcements"))):
    try:
        object_id = ObjectId(announcement_id)
    except Exception:
        raise HTTPException(400, "Invalid announcement ID.")
    result = await announcement_collection.update_one({"_id": object_id}, {"$set": {**payload.model_dump(), "updated_at": datetime.now(timezone.utc), "updated_by": user["_id"]}})
    if not result.matched_count: raise HTTPException(404, "Announcement not found.")
    await audit(user, "Updated announcement", announcement_id, payload.title)
    return {"success": True}

@router.delete("/announcements/{announcement_id}", dependencies=[Depends(csrf_guard)])
async def delete_announcement(announcement_id: str, user=Depends(require("manage_announcements"))):
    try:
        object_id = ObjectId(announcement_id)
    except Exception:
        raise HTTPException(400, "Invalid announcement ID.")
    result = await announcement_collection.update_one({"_id": object_id, "is_archived": {"$ne": True}}, {"$set": {"is_archived": True, "archived_at": datetime.now(timezone.utc), "archived_by": user["_id"]}})
    if not result.matched_count: raise HTTPException(404, "Announcement not found.")
    await audit(user, "Archived announcement", announcement_id)
    return {"success": True}
